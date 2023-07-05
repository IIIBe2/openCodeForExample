import os
from threading import Thread
import threading
from time import sleep
import time
from openpyxl import load_workbook
from subprocess import call

import requests 

class Erorr(Exception):
    def __init__(self, txt) -> None:
        self.text = txt

class Write_to_excel_array(Exception):
    def __init__(self) -> None:
        print("\n%2 != 2, error length... wait 1s...")
        sleep(1)

class WriteToFile:
    def __init__(self) -> None:
        self.SeekLine = 0
        self.Write_to_excel_array = []
        self.count_line_write = 0

    def Write_in_file(self, *, slot_data: object) -> None:
        self.SeekLine += len(slot_data)
        while slot_data:
            txt = slot_data.popitem()
            self.worksheet[txt[0]] = txt[1]
        self.File_save()

    def File_save(self) -> None:
        self.CopyFileToWriten = self.file
        self.CopyFileToWriten.save("DataTables/ExcelDataSave22023.xlsx")

    def Open_Excel(self, *, path: str) -> object:
        try:
            if self.worksheet != None: 
                raise Erorr("DONT_Empty_worksheet\nClear...")
        except AttributeError:
            self.file = load_workbook(path)
            self.worksheet = self.file.active
            return self.worksheet

    def MaxLine_reading_Excel(self) -> int:
        self.MaxLine = self.worksheet.max_row
        return self.MaxLine

    def Read_excel(self, *, Row: str, Line: str) -> str:
        return self.worksheet[Row + Line].value

    def Set_SeekLine(self, *, SeekLine: int):
        self.SeekLine = SeekLine

    def Write_wait(self) -> None:
        while self.MaxLine != self.SeekLine:
            for i in range(0, len(self.Write_to_excel_array), 2):
                self.worksheet[self.Write_to_excel_array[i]].value = self.Write_to_excel_array[i + 1]
                self.count_line_write += 1
            sleep(1)
    
class Inititation_Thread:
    def __init__(self, Threads_Num: int) -> None:
        self.th = []
        self.Threads_Num = Threads_Num

    def Thread_Create(self, *, classData: object) -> None:
        for i in range(0, self.Threads_Num):
            self.th.append(Thread(target = classData.Response_html(), args = (i), daemon = True))
    
    def Thread_start(self) -> None:
        for i in range(0, self.Threads_Num):
            self.th[i].start

    def Thread_live(self) -> None:
        self.Th_is_live = []
        for Live_Th in range(0,  self.Threads_Num):
            if self.th[Live_Th].is_alive() is False:
                self.Th_is_live.append(Live_Th)
    
    def Thread_live_print(self) -> None:
        if not self.Th_is_live:
            print("List empty")
        else:
            print("List:", str(self.Th_is_live))

class Console_write:
    def __init__(self) -> None:
        pass

    def Console_restart_tor(self, *, password: str):
        call('echo {} | sudo -S {}'.format(password, 'sudo service tor restart'), shell=True)
        #sleep(0.05)
        sleep(0.2)

    def Console_status_tor(self) -> None:
        os.system("service tor status")

class Collect_Data:
    def __init__(self, maxLine, WriteToFileClass) -> None:
        self.list_url = {}
        self.numeric_to_write = 0
        self.list_port = list(reversed(range(1, 1000)))
        self.slot = list(reversed(range(1, maxLine)))
        self.SeekLine = 1
        self.data_to_write = {}
        self.maxLine = maxLine
        self.Read_excel = WriteToFileClass.Read_excel
        self.list_html = {}
        self.list_Name_Slot = {}
        self.response = [[]] * 10
        self.local_url= [[]] * 10
        self.Standart_url = "https://steamcommunity.com/market/itemordershistogram?country=RU&language=russian&currency=1&item_nameid="
        self.Update_list_url(amount=0)
        

    def Update_list_url(self, *, amount: int):
        if amount != 0 and (self.maxLine - amount - self.SeekLine) > 0:
            for i in reversed(range(self.SeekLine, amount + 1)):
                try:
                    if self.Read_excel(Row="E", Line=str(i)) == None:
                        self.list_url[self.Read_excel(Row="A", Line=str(i))] = self.Standart_url + self.Read_excel(Row="D", Line=str(i))
                        self.list_Name_Slot[self.Read_excel(Row="A", Line=str(i))] = "E" + str(i)
                except TypeError:
                    ...
        elif amount == 0 and (self.maxLine - amount - self.SeekLine) > 0:
            for i in reversed(range(self.SeekLine, self.maxLine)):
                try:
                    d = self.Read_excel(Row="E", Line=str(i))
                    if d == None:
                        self.list_url[self.Read_excel(Row="A", Line=str(i))] = self.Standart_url + self.Read_excel(Row="D", Line=str(i))
                        self.list_Name_Slot[self.Read_excel(Row="A", Line=str(i))] = "E" + str(i)
                except TypeError:
                    ...
        else:
            raise Erorr("Not correctly amount")
        
    def Response_html(self, *, th: int):
        while True:
            if len(self.list_url) > 0:
                try:
                    self.local_port = self.Get_Port()
                    while len(self.list_url) > 0:
                        self.local_url[th] = self.Get_Url()
                        self.response[th] = self.Create_session(port=self.local_port).get(self.local_url[th][1], timeout=3)
                        if self.response[th].text == "null":
                            while self.response[th].text == "null":
                                self.response[th] = self.Create_session(port=self.local_port).get(self.local_url[th][1], timeout=3)
                                self.local_port = self.Get_Port()
                        self.list_html[self.local_url[th][0]] = self.response[th].text
                        #print(str(self.local_port )+ "url_unloc:", self.response[th].url + "\n" + str(self.local_port )+ "url_local:", self.local_url[th][1])
                except (ConnectionRefusedError, requests.exceptions.ConnectTimeout, requests.ConnectionError, requests.ReadTimeout):
                    ...
                    #print("bad port")
    
    def pri(self):
        print(self.list_url)

    def Create_session(self, *, port) -> object:
        self.sessionReade = requests.session()
        self.sessionReade.proxies = {}
        if port < 10:
            self.sessionReade.proxies['http']='socks5://127.0.0.1:900' + str(port)
            self.sessionReade.proxies['https']='socks5://127.0.0.1:900' + str(port)
        elif port == 51:
            pass
        elif port < 100:
            self.sessionReade.proxies['http']='socks5://127.0.0.1:90' + str(port)
            self.sessionReade.proxies['https']='socks5://127.0.0.1:90' + str(port)
        elif port < 1000:
            self.sessionReade.proxies['http']='socks5://127.0.0.1:9' + str(port)
            self.sessionReade.proxies['https']='socks5://127.0.0.1:9' + str(port)
        return self.sessionReade
		
    def Collect_data_dict(self, coordinate: str, text: str) -> None:
        self.data_to_write[coordinate] = text

    def add_to_write(self) -> None:
        self.numeric_to_write += 1

    def get_numeric_to_write(self) -> int:
        return self.numeric_to_write

    def Get_Collect_data_dict(self)-> object:
        return self.data_to_write
    
    def Clear_Collect_data_dict(self)-> None:
        self.data_to_write = {}

    def Get_html(self):
        if self.list_html:
            return self.list_html.popitem()

    def Get_Url(self) -> object:
        if len(self.list_url)> 0:
            return self.list_url.popitem()
    
    def Restart_Port(self):
        self.list_port = list(reversed(range(1, 1000)))
        call('echo {} | sudo -S {}'.format("123123", 'sudo service tor restart'), shell=True)
        sleep(0.2)

    def Get_Port(self) -> int:
        if len(self.list_port) > 0:
            return self.list_port.pop()
        else: 
            self.Restart_Port()
            return self.list_port.pop()

class Search_data:
    def __init__(self, Data_class) -> None:
        self.list_html_local = Data_class.list_html
        self.slot = Data_class.slot
        self.list_Name_Slot = Data_class.list_Name_Slot
        self.TempList = {"indexPriceS": 0, "indexPriceE": 0 }

    def Search_Price(self, *, Data_class) -> None:
        self.list_html_local = Data_class.Get_html()
        self.TempList["indexPriceS"] = 0
        try:
            if len(self.list_html_local) > 0:
                self.page_html = self.list_html_local
                for i in range(0, 11):
                    self.TempList["indexPriceS"] = self.page_html[1].find("class=", self.TempList["indexPriceS"] + 2)
                self.TempList["indexPriceE"] = self.page_html[1].find("td", self.TempList["indexPriceS"])
                if self.TempList["indexPriceS"] != -1 and self.TempList["indexPriceE"] != None and self.TempList["indexPriceE"] != -1:
                    if (self.TempList["indexPriceE"] - 11 - self.TempList["indexPriceS"] + 3) < 15:
                        print("-", self.page_html[0].ljust(50), "-|-" , self.page_html[1][self.TempList["indexPriceS"]+11:self.TempList["indexPriceE"]-3])
                        Data_class.Collect_data_dict(self.list_Name_Slot[self.page_html[0]], self.page_html[1][self.TempList["indexPriceS"]+11:self.TempList["indexPriceE"]-3])
                        Data_class.add_to_write()
                    else:
                        print("-", self.page_html[0].ljust(50), "-|-" , "$0.03   defoult")
                        Data_class.Collect_data_dict(self.list_Name_Slot[self.page_html[0]], "$0.03")
                        Data_class.add_to_write()
        except TypeError:
            ...

class MyThreads(Thread):
    def __init__(self, num):
        Thread.__init__(self)
        self.name = str(num)
        self.num = num

    def run(self):
        Main_Select.Response_html(th=self.num)

class MyThreads_control(Thread):
    def __init__(self, num, *, Collect_data, Writing_class):
        self.Main_Select = Collect_data
        self.Writing_class = Writing_class
        self.Write_N = 0
        
        Thread.__init__(self)
        self.name = "controle"
        self.maxLine = Writing_class.MaxLine_reading_Excel()

    def update_status_list_urls(self) -> int:
        return self.maxLine - self.Writing_class.SeekLine


    def run(self):
        toc = time.perf_counter()
        Support_Select = Search_data(self.Main_Select)
        while self.update_status_list_urls() > 0:
            self.update_status_list_urls
            Support_Select.Search_Price(Data_class=self.Main_Select)
            if self.Write_N < int(self.Main_Select.get_numeric_to_write() / 100):
                tic = time.perf_counter()
                h = int(((tic - toc) / 60) / 60)
                m = int(((tic - toc) - (h * 3600)) / 60)
                s = int((tic - toc) % 60)
                self.Write_N = int(self.Main_Select.get_numeric_to_write() / 100)
                print(f"\nВычисление заняло {h} ч {m} м {s} с, сохранение {self.Write_N * 100}...")
                #print(f"Вычисление заняло {tic - toc:0.2f} секунд, сохранение...")
                self.Writing_class.Write_in_file(slot_data=self.Main_Select.Get_Collect_data_dict())
                self.Main_Select.Clear_Collect_data_dict()
            sleep(1)

if __name__ == "__main__":
    Writing_while = WriteToFile()
    Writing_while.Open_Excel(path="DataTables/ExcelDataSave22023.xlsx")
    maxLine = Writing_while.MaxLine_reading_Excel()
    Console = Console_write()
    Console.Console_restart_tor(password=123123)
    Console.Console_status_tor()
    Main_Select = Collect_Data(maxLine, Writing_while)
    th = []
    for i in range(0, 10):
        th.append(MyThreads(i))
    th_controle = MyThreads_control(1, Collect_data=Main_Select, Writing_class=Writing_while)
    """th1.append(Thread(target = MyThreads, args=(1)))
    th2.append(Thread(target = MyThreads, args=(2))) """
    for i in range(0, 10):
        th[i].start()
    th_controle.start()
    for i in range(0, 10):
        th[i].join()
    th_controle.join()